from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import re
import io

app = Flask(__name__)
CORS(app, origins=["https://real-estate-webpage-black.vercel.app", "http://localhost:3000"])

DB_FILE = "leads.db"
HEADERS = ["S.No", "Date & Time", "Client Name", "Requirement", "Budget (Lakhs)", "Mobile Number", "Sales Executive"]


def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datetime TEXT NOT NULL,
            client_name TEXT NOT NULL,
            requirement TEXT NOT NULL,
            budget REAL NOT NULL,
            mobile TEXT NOT NULL,
            sales_executive TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


def get_all_leads():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT id, datetime, client_name, requirement, budget, mobile, sales_executive FROM leads ORDER BY id")
    rows = c.fetchall()
    conn.close()
    return rows


# Initialize DB at module load so gunicorn also sets it up
init_db()


@app.route("/api/submit", methods=["POST"])
def submit_lead():
    data = request.get_json()

    required = ["clientName", "requirement", "budget", "mobile", "salesExecutive"]
    for field in required:
        if not data.get(field):
            return jsonify({"success": False, "message": f"Missing field: {field}"}), 400

    mobile = str(data["mobile"]).strip()
    if not re.fullmatch(r"[6-9]\d{9}", mobile):
        return jsonify({"success": False, "message": "Invalid mobile number. Must be a 10-digit Indian mobile number."}), 400

    try:
        budget = float(data["budget"])
        if budget <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"success": False, "message": "Budget must be a positive number."}), 400

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        "INSERT INTO leads (datetime, client_name, requirement, budget, mobile, sales_executive) VALUES (?, ?, ?, ?, ?, ?)",
        (
            datetime.now().strftime("%d-%m-%Y %H:%M"),
            data["clientName"],
            data["requirement"],
            budget,
            mobile,
            data["salesExecutive"],
        ),
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Lead submitted successfully!"})


@app.route("/api/leads", methods=["GET"])
def get_leads():
    rows = get_all_leads()
    leads = [dict(zip(HEADERS, row)) for row in rows]
    return jsonify(leads)


@app.route("/api/export", methods=["GET"])
def export_leads():
    rows = get_all_leads()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1A3C5E")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [6, 22, 22, 16, 18, 18, 20]
    for i, (header, width) in enumerate(zip(HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    even_fill = PatternFill("solid", start_color="EAF1FB")
    odd_fill = PatternFill("solid", start_color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    for idx, row_data in enumerate(rows, start=1):
        fill = even_fill if idx % 2 == 0 else odd_fill
        row_num = idx + 1
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = border
            cell.fill = fill
            cell.alignment = align_center if col in [1, 2, 4, 5, 7] else align_left
        ws.row_dimensions[row_num].height = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"basil_flora_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(debug=True, port=5000)
